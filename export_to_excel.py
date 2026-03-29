"""Export speed test and traceroute data from SQLite to a formatted Excel file."""

import datetime
import os
import sqlite3
import sys
from openpyxl import Workbook
from openpyxl.styles import numbers

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(SCRIPT_DIR, "speedtest.db")
EXCEL_PATH = os.path.join(SCRIPT_DIR, "SpeedTestResults.xlsx")
EXCEL_EPOCH = datetime.datetime(1899, 12, 30)


def serial_to_datetime(serial):
    """Convert an Excel serial date number back to a Python datetime."""
    try:
        serial = float(serial)
        return EXCEL_EPOCH + datetime.timedelta(days=serial)
    except (TypeError, ValueError):
        return None


def export():
    if not os.path.exists(DB_PATH):
        print("No database found. Run the agent first.")
        sys.exit(1)

    conn = sqlite3.connect(DB_PATH)
    wb = Workbook()

    # --- Speed Tests sheet ---
    ws = wb.active
    ws.title = "Speed Tests"
    ws.append(["Timestamp", "Download (Mbps)", "Upload (Mbps)", "Ping (ms)", "Server", "Host", "ISP", "Error"])

    rows = conn.execute(
        "SELECT timestamp, download_mbps, upload_mbps, ping_ms, server_name, server_host, isp, error "
        "FROM speed_tests ORDER BY timestamp"
    ).fetchall()

    for row in rows:
        dt = serial_to_datetime(row[0])
        ws.append([dt, row[1], row[2], row[3], row[4], row[5], row[6], row[7]])

    # Format timestamp column as date/time
    for cell in ws["A"][1:]:
        cell.number_format = "M/D/YYYY H:MM:SS AM/PM"

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # --- Traceroutes sheet ---
    ws2 = wb.create_sheet("Traceroutes")
    ws2.append(["Timestamp", "Target", "Hop #", "Hop IP", "RTT (ms)", "Error"])

    rows = conn.execute(
        "SELECT timestamp, target, hop_number, hop_ip, rtt_ms, error "
        "FROM traceroutes ORDER BY timestamp, hop_number"
    ).fetchall()

    for row in rows:
        dt = serial_to_datetime(row[0])
        ws2.append([dt, row[1], row[2], row[3], row[4], row[5]])

    for cell in ws2["A"][1:]:
        cell.number_format = "M/D/YYYY H:MM:SS AM/PM"

    for col in ws2.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    conn.close()

    wb.save(EXCEL_PATH)
    print(f"Exported to: {EXCEL_PATH}")
    print(f"  Speed tests:  {len(ws['A']) - 1} rows")
    print(f"  Traceroutes:  {len(ws2['A']) - 1} rows")


if __name__ == "__main__":
    export()
